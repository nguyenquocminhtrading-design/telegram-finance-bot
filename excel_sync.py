import os
import openpyxl
from datetime import date

EXPENSE_FILE = "My expenses.xlsx"
PORTFOLIO_FILE = "My portfolio.xlsm"

def _next_empty_row(ws):
    """Find the first empty row by scanning column A from the top."""
    for row in range(1, ws.max_row + 2):
        if ws.cell(row=row, column=1).value is None:
            return row
    return ws.max_row + 1

def sync_expense_to_excel(transaction_data):
    """
    transaction_data: dict with keys:
    date, amount, category, description, bank_account
    """
    try:
        if not os.path.exists(EXPENSE_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Expenses"
            ws.append(["Date", "Amount", "Category", "Description", "Bank Account"])
        else:
            wb = openpyxl.load_workbook(EXPENSE_FILE)
            ws = wb.active

        next_row = _next_empty_row(ws)
        row_data = [
            transaction_data.get("date", date.today().isoformat()),
            transaction_data.get("amount", 0),
            transaction_data.get("category", "other"),
            transaction_data.get("description", ""),
            transaction_data.get("bank_account", "")
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=next_row, column=col_idx, value=value)

        wb.save(EXPENSE_FILE)
    except Exception as e:
        print(f"Error syncing to expense excel: {e}")

def sync_asset_to_portfolio(asset_data, is_buy=True):
    """
    asset_data: dict with keys:
    date, name, value, note
    """
    try:
        if not os.path.exists(PORTFOLIO_FILE):
            print(f"Portfolio file {PORTFOLIO_FILE} not found.")
            return

        # keep_vba=True to preserve macros in .xlsm
        wb = openpyxl.load_workbook(PORTFOLIO_FILE, keep_vba=True)
        if "Transaction" in wb.sheetnames:
            ws = wb["Transaction"]
        else:
            # Fallback to active sheet if Transaction doesn't exist
            ws = wb.active
            
        trans_type = "Mua" if is_buy else "Bán"
        net_flow = -asset_data.get("value", 0) if is_buy else asset_data.get("value", 0)

        next_row = _next_empty_row(ws)
        row_data = [
            asset_data.get("date", date.today().isoformat()),
            trans_type,
            asset_data.get("name", ""),
            asset_data.get("value", 0),
            0, # Phí GD
            0, # Thuế bán
            net_flow, # Dòng tiền ròng
            asset_data.get("note", "")
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=next_row, column=col_idx, value=value)
        
        wb.save(PORTFOLIO_FILE)
    except Exception as e:
        print(f"Error syncing to portfolio excel: {e}")
